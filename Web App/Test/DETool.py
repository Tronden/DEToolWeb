import threading
import webbrowser
import os
import sys
import json
import io
import time
import requests
import pandas as pd
import numpy as np
from flask import Flask, send_from_directory, request, jsonify, make_response, Response
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import pystray
from pystray import Menu, MenuItem
from PIL import Image, ImageDraw
import subprocess
import concurrent.futures
import datetime
import logging
from logging.handlers import RotatingFileHandler
import csv
from threading import Lock

###############################################################################
# GLOBAL CONCURRENCY LOCK
###############################################################################
global_lock = Lock()

###############################################################################
# LOGGING SETUP (rotating)
###############################################################################
def get_logs_folder():
    p = os.path.join(os.path.expanduser("~"), "Documents", "DETool", "Logs")
    os.makedirs(p, exist_ok=True)
    return p

LOGS_FOLDER = get_logs_folder()

def make_rotating_handler(logname):
    return RotatingFileHandler(
        os.path.join(LOGS_FOLDER, logname),
        mode='a',
        maxBytes=5_000_000,  # 5 MB
        backupCount=5,
        encoding=None,
        delay=0
    )

python_logger = logging.getLogger("python_exec")
python_logger.setLevel(logging.INFO)
py_handler = make_rotating_handler("python_execution.log")
py_formatter = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")
py_handler.setFormatter(py_formatter)
python_logger.addHandler(py_handler)

script_logger = logging.getLogger("script_exec")
script_logger.setLevel(logging.INFO)
script_handler = make_rotating_handler("script_execution.log")
script_handler.setFormatter(py_formatter)
script_logger.addHandler(script_handler)

user_logger = logging.getLogger("user_interactions")
user_logger.setLevel(logging.INFO)
user_handler = make_rotating_handler("user_interactions.log")
user_handler.setFormatter(py_formatter)
user_logger.addHandler(user_handler)

python_logger.info("DETool server starting up...")

###############################################################################
# FLASK APP + CONSTANTS
###############################################################################
UI_PORT = 8080
DATA_DIR = "data"

EXTERNAL_TAGLIST_URL = "http://localhost:61185/taglist"
EXTERNAL_VALUES_URL = "http://localhost:61185/values"

app = Flask(__name__, static_folder=DATA_DIR)

###############################################################################
# GLOBAL IN-MEMORY
###############################################################################
RAW_TABLE = None
WORKING_TABLE = None
TAGLIST_CACHE = None
LAST_SETTINGS = None
TAG_COVERAGE = {}

# CHANGED: Add a global signature to track changes to RAW_TABLE
RAW_TABLE_SIGNATURE = None

###############################################################################
# PATH HELPERS
###############################################################################
def get_base_folder():
    return os.path.join(os.path.expanduser("~"), "Documents", "DETool")

def get_cache_folder():
    p = os.path.join(get_base_folder(), "Cache")
    os.makedirs(p, exist_ok=True)
    return p

def get_settings_folder():
    p = os.path.join(get_base_folder(), "Settings")
    os.makedirs(p, exist_ok=True)
    return p

def get_site_settings_path():
    return os.path.join(get_settings_folder(), "SiteSettings.json")

def get_tag_settings_path():
    return os.path.join(get_settings_folder(), "TagSettings.json")

def get_taglist_cache_path():
    return os.path.join(get_cache_folder(), "Taglist.json")

def get_working_table_cache_path():
    return os.path.join(get_cache_folder(), "WorkingTable.json")

def get_raw_table_cache_path():
    return os.path.join(get_cache_folder(), "RawTable.json")

def get_tag_coverage_cache_path():
    return os.path.join(get_cache_folder(), "TagCoverage.json")

def fmt_timestamp(dt):
    return dt.strftime("%d/%m/%Y %H:%M:%S")

###############################################################################
# ATOMIC JSON & DATAFRAME SAVE/LOAD
###############################################################################
def atomic_write_json(path, data):
    tmp_path = path + ".tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(data, f)
    os.replace(tmp_path, path)

def safe_load_json(path, default=None):
    if default is None:
        default = {}
    if not os.path.exists(path):
        return default
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        python_logger.error(f"Error loading JSON from {path}: {e}")
        return default

def save_df_to_json(df, path):
    if df is None or df.empty:
        payload = {"columns": [], "data": []}
    else:
        payload = {
            "columns": df.columns.tolist(),
            "data": df.values.tolist()
        }
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(payload, f)
    os.replace(tmp, path)

def load_df_from_json(path):
    if not os.path.exists(path):
        return pd.DataFrame()
    try:
        with open(path, "r", encoding="utf-8") as f:
            payload = json.load(f)
        cols = payload.get("columns", [])
        data = payload.get("data", [])
        if not cols or not data:
            return pd.DataFrame(columns=cols)
        return pd.DataFrame(data, columns=cols)
    except Exception as e:
        python_logger.error(f"Error loading DF from JSON {path}: {e}")
        return pd.DataFrame()

###############################################################################
# LOAD/SAVE RAW & WORKING, TAG_COVERAGE
###############################################################################
def save_raw_table_cache():
    global RAW_TABLE
    if RAW_TABLE is not None:
        p = get_raw_table_cache_path()
        try:
            save_df_to_json(RAW_TABLE, p)
            python_logger.info("RAW_TABLE cached successfully.")
        except Exception as e:
            python_logger.error(f"Error caching RAW_TABLE: {e}")

def load_raw_table_cache():
    global RAW_TABLE
    p = get_raw_table_cache_path()
    if os.path.exists(p):
        try:
            df = load_df_from_json(p)
            if not df.empty:
                python_logger.info("Loaded RAW_TABLE from JSON cache.")
                return df
        except Exception as e:
            python_logger.error(f"Error loading RAW_TABLE: {e}")
    # if missing or error, return empty
    return pd.DataFrame(columns=["NumericTimestamp", "Timestamp"])

def save_working_table_cache():
    global WORKING_TABLE
    if WORKING_TABLE is not None:
        p = get_working_table_cache_path()
        try:
            save_df_to_json(WORKING_TABLE, p)
            python_logger.info("WORKING_TABLE cached successfully.")
        except Exception as e:
            python_logger.error(f"Error caching WORKING_TABLE: {e}")

def load_working_table_cache():
    p = get_working_table_cache_path()
    if os.path.exists(p):
        try:
            df = load_df_from_json(p)
            python_logger.info("Loaded WORKING_TABLE from JSON cache.")
            return df
        except Exception as e:
            python_logger.error(f"Error loading WORKING_TABLE: {e}")
    return None

def save_tag_coverage():
    global TAG_COVERAGE
    p = get_tag_coverage_cache_path()
    try:
        atomic_write_json(p, TAG_COVERAGE)
        python_logger.info("TAG_COVERAGE cached successfully.")
    except Exception as e:
        python_logger.error(f"Error caching TAG_COVERAGE: {e}")

def load_tag_coverage():
    p = get_tag_coverage_cache_path()
    if os.path.exists(p):
        try:
            with open(p, "r", encoding="utf-8") as f:
                coverage = json.load(f)
            python_logger.info("Loaded TAG_COVERAGE from JSON cache.")
            return coverage
        except Exception as e:
            python_logger.error(f"Error loading TAG_COVERAGE: {e}")
    return {}

###############################################################################
# SIGNATURE HELPER (to detect if RAW_TABLE changed)
###############################################################################
def get_raw_table_signature(df: pd.DataFrame):
    """
    Returns a simple signature tuple: (row_count, sorted_columns, last_numeric_timestamp)
    to detect data changes in RAW_TABLE.
    """
    if df is None or df.empty:
        return (0, tuple(), None)
    row_count = df.shape[0]
    col_tuple = tuple(sorted(df.columns))
    # If sorted by NumericTimestamp, this is the last row's numeric TS
    last_ts = df["NumericTimestamp"].iloc[-1] if "NumericTimestamp" in df.columns else None
    return (row_count, col_tuple, last_ts)

###############################################################################
# INTERVAL HELPERS
###############################################################################
def union_intervals(intervals):
    if not intervals:
        return []
    sorted_iv = sorted(intervals, key=lambda x: x[0])
    merged = [sorted_iv[0]]
    for i in range(1, len(sorted_iv)):
        cur = sorted_iv[i]
        last = merged[-1]
        if cur[0] <= last[1]:
            merged[-1] = (last[0], max(last[1], cur[1]))
        else:
            merged.append(cur)
    return merged

def remove_tag_coverage(tag):
    global TAG_COVERAGE, RAW_TABLE
    if tag in TAG_COVERAGE:
        del TAG_COVERAGE[tag]
    if RAW_TABLE is not None and not RAW_TABLE.empty and tag in RAW_TABLE.columns:
        RAW_TABLE.drop(columns=[tag], inplace=True, errors="ignore")

###############################################################################
# LOGGING ENDPOINT
###############################################################################
@app.route("/log_event", methods=["POST"])
def log_event():
    try:
        data = request.get_json()
        msg_type = data.get("type", "script").lower()
        message = data.get("message", "")
        if msg_type == "user":
            user_logger.info(message)
        else:
            script_logger.info(message)
        return jsonify({"status": "logged"})
    except Exception as e:
        python_logger.error(f"Error in /log_event: {e}")
        return jsonify({"error": str(e)}), 500

###############################################################################
# STATIC FILES
###############################################################################
@app.route("/")
def root():
    return send_from_directory(DATA_DIR, "index.html")

@app.route("/<path:fname>")
def serve_static(fname):
    return send_from_directory(DATA_DIR, fname)

###############################################################################
# TAGLIST
###############################################################################
@app.route("/taglist")
def taglist():
    global TAGLIST_CACHE
    cachep = get_taglist_cache_path()
    refresh = request.args.get("refresh", "false").lower() in ["true", "1"]

    if not refresh and TAGLIST_CACHE:
        python_logger.info("Returning taglist from in-memory cache.")
        return jsonify(TAGLIST_CACHE)

    if not refresh and os.path.exists(cachep):
        try:
            with open(cachep, "r") as f:
                data = json.load(f)
                TAGLIST_CACHE = data
                python_logger.info("Returning taglist from disk cache.")
                return jsonify(data)
        except Exception as e:
            python_logger.error(f"Error reading Taglist cache: {e}")

    try:
        python_logger.info("Fetching new taglist from external source...")
        r = requests.get(EXTERNAL_TAGLIST_URL, timeout=15)
        r.raise_for_status()
        data = r.json()
        atomic_write_json(cachep, data)
        TAGLIST_CACHE = data
        python_logger.info("Taglist fetched and cached.")
        return jsonify(data)
    except requests.exceptions.ConnectionError as ce:
        python_logger.error(f"Failed to connect to external taglist: {ce}")
        if TAGLIST_CACHE:
            python_logger.info("Returning in-memory taglist due to error.")
            return jsonify(TAGLIST_CACHE)
        return jsonify({"error": "No connection & no cached taglist"}), 503
    except Exception as e:
        python_logger.error(f"Failed to fetch external taglist: {e}")
        if TAGLIST_CACHE:
            return jsonify(TAGLIST_CACHE)
        return jsonify([])

###############################################################################
# FETCH SINGLE TAG
###############################################################################
def fetch_values(tag, st, en):
    try:
        r = requests.get(
            EXTERNAL_VALUES_URL,
            params={"tag": tag, "startDateUnixSeconds": st, "endDateUnixSeconds": en},
            timeout=15
        )
        r.raise_for_status()
        return tag, r.json()
    except Exception as e:
        python_logger.error(f"fetch_values failed for {tag} ({st}-{en}): {e}")
        return tag, []

###############################################################################
# MERGE => RAW_TABLE
###############################################################################
def merge_new_data_into_raw_table(df_new):
    """
    Merges df_new (with columns ["NumericTimestamp","Timestamp","tagName"]) into RAW_TABLE
    overwriting old data if there's overlap in the same timestamps.
    """
    global RAW_TABLE
    if RAW_TABLE is None or RAW_TABLE.empty:
        RAW_TABLE = df_new.sort_values("NumericTimestamp").reset_index(drop=True)
        return

    combined = pd.merge(
        RAW_TABLE,
        df_new,
        on="NumericTimestamp",
        how="outer",
        suffixes=("", "_new")
    )

    for c in df_new.columns:
        if c in ("Timestamp", "NumericTimestamp"):
            continue
        new_col = c + "_new"
        if new_col in combined.columns:
            # Overwrite old with new if new is not NaN
            combined[c] = combined[c].where(combined[new_col].isna(), combined[new_col])
            combined.drop(columns=[new_col], inplace=True)

    if "Timestamp_new" in combined.columns:
        combined.drop(columns=["Timestamp_new"], inplace=True)

    # Rebuild consistent "Timestamp" from NumericTimestamp
    combined["Timestamp"] = combined["NumericTimestamp"].apply(
        lambda x: fmt_timestamp(pd.to_datetime(x, unit="ms"))
    )
    combined.sort_values("NumericTimestamp", inplace=True)
    combined.reset_index(drop=True, inplace=True)
    RAW_TABLE = combined

###############################################################################
# FORWARD-FILL
###############################################################################
def build_filled_df_from_raw_table():
    global RAW_TABLE
    if RAW_TABLE is None or RAW_TABLE.empty:
        return None
    tgSetData = safe_load_json(get_tag_settings_path(), {
        "scale_factors": {}, "error_value": {}, "max_decimal": {}
    })
    err_vals = tgSetData.get("error_value", {})
    df_filled = RAW_TABLE.copy()
    for c in df_filled.columns:
        if c in ["Timestamp", "NumericTimestamp"]:
            continue
        vals = pd.to_numeric(df_filled[c], errors="coerce")
        if c in err_vals:
            try:
                badv = float(err_vals[c])
                vals = vals.mask(vals == badv, np.nan)
            except:
                pass
        df_filled[c] = vals
    df_filled = df_filled.ffill()
    return df_filled

###############################################################################
# BUILD WORKING TABLE
###############################################################################
def build_working_table(offset_hours=0, forward_fill=False):
    global WORKING_TABLE, LAST_SETTINGS, RAW_TABLE
    if RAW_TABLE is None or RAW_TABLE.empty:
        WORKING_TABLE = None
        return

    tgSetData = safe_load_json(get_tag_settings_path(), {
        "scale_factors": {}, "error_value": {}, "max_decimal": {}
    })
    err_vals = tgSetData.get("error_value", {})

    if forward_fill:
        base_df = build_filled_df_from_raw_table()
    else:
        base_df = RAW_TABLE.copy()
        for c in base_df.columns:
            if c in ["Timestamp", "NumericTimestamp"]:
                continue
            vals = pd.to_numeric(base_df[c], errors="coerce")
            if c in err_vals:
                try:
                    badv = float(err_vals[c])
                    vals = vals.mask(vals == badv, np.nan)
                except:
                    pass
            base_df[c] = vals

    if base_df is None or base_df.empty:
        WORKING_TABLE = None
        return

    offMs = int(offset_hours * 3600000)
    wdf = base_df.copy()
    wdf["NumericTimestamp"] = wdf["NumericTimestamp"] + offMs
    wdf["Timestamp"] = wdf["NumericTimestamp"].apply(
        lambda x: fmt_timestamp(pd.to_datetime(x, unit="ms"))
    )

    sf = tgSetData.get("scale_factors", {})
    dec = tgSetData.get("max_decimal", {})

    for c in wdf.columns:
        if c in ["Timestamp", "NumericTimestamp"]:
            continue
        vals = pd.to_numeric(wdf[c], errors="coerce")
        sc_factor = float(sf.get(c, 1))
        decimals = int(dec.get(c, 2))
        vals = (vals * sc_factor).round(decimals)
        wdf[c] = vals

    WORKING_TABLE = wdf.copy()

###############################################################################
# PARTIAL FETCH => RAW_TABLE
###############################################################################
@app.route("/fetch_data", methods=["POST"])
def fetch_data_endpoint():
    """
    Fetch new data only for time intervals not yet covered by TAG_COVERAGE.
    Merge into RAW_TABLE if new data is received.
    Return { newData: true/false, redrawNeeded: true/false } accordingly.
    """
    global RAW_TABLE, TAG_COVERAGE, RAW_TABLE_SIGNATURE
    req = request.get_json()
    if not req:
        return jsonify({"error": "Invalid JSON"}), 400

    tags = req.get("tags", [])
    st = req.get("startDateUnixSeconds")
    en = req.get("endDateUnixSeconds")
    autoRefresh = bool(req.get("autoRefresh", False))
    user_logger.info(f"/fetch_data with tags={tags}, st={st}, en={en}, autoRefresh={autoRefresh}")
    if not tags or st is None or en is None:
        return jsonify({"error": "Missing fields"}), 400

    data_changed = False
    old_signature = RAW_TABLE_SIGNATURE

    with global_lock:
        # Drop any removed tags from coverage
        old_tags = set(TAG_COVERAGE.keys())
        new_tags = set(tags)
        removed = old_tags - new_tags
        for rt in removed:
            remove_tag_coverage(rt)

        if RAW_TABLE is None:
            RAW_TABLE = pd.DataFrame(columns=["NumericTimestamp","Timestamp"])

        futs = []
        executor_size = min(len(tags), 4)
        with concurrent.futures.ThreadPoolExecutor(max_workers=executor_size) as exe:
            for tg in tags:
                if tg not in TAG_COVERAGE:
                    TAG_COVERAGE[tg] = []

                coverage_list = union_intervals(TAG_COVERAGE[tg])
                needed = (st, en)
                missing = []
                cS = needed[0]
                cE = needed[1]

                # Figure out sub-intervals not covered
                for (cvS, cvE) in coverage_list:
                    # if coverage doesn't overlap
                    if cvE < cS or cvS > cE:
                        continue
                    # partial coverage
                    if cvS > cS:
                        missing.append((cS, min(cvS, cE)))
                    if cvE > cS:
                        cS = max(cS, cvE)
                    if cS > cE:
                        break

                if cS < cE:
                    missing.append((cS, cE))

                # Submit fetch tasks for missing intervals
                for (miS, miE) in missing:
                    if miE <= miS:
                        continue
                    fut = exe.submit(fetch_values, tg, miS, miE)
                    futs.append((fut, tg, miS, miE))

        # Process fetch results
        for (fut, tg, fs, fe) in futs:
            try:
                tagFetched, arr = fut.result()
                if arr:
                    df = pd.DataFrame(arr)
                else:
                    df = pd.DataFrame(columns=["Date", "Value"])
                if df.empty:
                    continue

                df["Value"] = pd.to_numeric(df["Value"], errors="coerce")
                df.replace([np.inf, -np.inf], np.nan, inplace=True)
                df["Timestamp"] = pd.to_datetime(df["Date"], errors="coerce")
                # NumericTimestamp in ms
                df["NumericTimestamp"] = (df["Timestamp"].astype(np.int64) // 1_000_000)
                df.sort_values("Timestamp", inplace=True)
                df_ren = df[["NumericTimestamp", "Timestamp", "Value"]].rename(columns={"Value": tg})

                merge_new_data_into_raw_table(df_ren)
                TAG_COVERAGE[tg].append((fs, fe))
                TAG_COVERAGE[tg] = union_intervals(TAG_COVERAGE[tg])

            except Exception as e:
                python_logger.error(f"Error partial fetching {tg} {fs}..{fe} => {e}")

        # Compare new signature to see if RAW_TABLE changed
        new_signature = get_raw_table_signature(RAW_TABLE)
        data_changed = (new_signature != old_signature)
        if data_changed:
            RAW_TABLE_SIGNATURE = new_signature
            save_raw_table_cache()
            save_tag_coverage()

    # If data didn't change, no need to rebuild on front end
    return jsonify({"status": "ok", "newData": data_changed, "redrawNeeded": data_changed})

###############################################################################
# BUILD WORKING_TABLE => FRONT-END
###############################################################################
@app.route("/build_working_table", methods=["POST"])
def api_build_working_table():
    global LAST_SETTINGS, WORKING_TABLE, RAW_TABLE
    if RAW_TABLE is None or RAW_TABLE.empty:
        return jsonify({"data": [], "redrawNeeded": False})

    req = request.get_json()
    dataOffset = float(req.get("dataOffset", 0))
    forwardFill = bool(req.get("forwardFill", False))
    if LAST_SETTINGS is None:
        LAST_SETTINGS = {}

    need_rebuild = False

    # Compare last known settings
    if (LAST_SETTINGS.get("dataOffset") != dataOffset or
        LAST_SETTINGS.get("forwardFill") != forwardFill):
        need_rebuild = True

    # If the RAW_TABLE row count changed, or columns changed, etc.
    old_len = len(WORKING_TABLE) if WORKING_TABLE is not None else 0
    raw_len = len(RAW_TABLE)
    if raw_len != old_len:
        need_rebuild = True

    if need_rebuild:
        user_logger.info(f"Rebuilding WORKING_TABLE with offset={dataOffset}, ff={forwardFill}")
        with global_lock:
            build_working_table(offset_hours=dataOffset, forward_fill=forwardFill)
            LAST_SETTINGS = {"dataOffset": dataOffset, "forwardFill": forwardFill}
            save_working_table_cache()
    else:
        python_logger.info("No rebuild needed for WORKING_TABLE.")

    if WORKING_TABLE is None:
        return jsonify({"data": [], "redrawNeeded": need_rebuild})

    df_safe = WORKING_TABLE.replace([np.inf, -np.inf, np.nan], None)
    return jsonify({"data": df_safe.to_dict(orient="records"), "redrawNeeded": need_rebuild})

###############################################################################
# EXPORT EXCEL
###############################################################################
@app.route("/export_excel", methods=["POST"])
def export_excel():
    global WORKING_TABLE
    try:
        req = request.get_json()
        start_ms = req.get("startDateUnixMillis")
        end_ms = req.get("endDateUnixMillis")
        bname = req.get("bargeName", "UnknownBarge")
        fnum = req.get("fhNumber", "0000")

        if WORKING_TABLE is None or WORKING_TABLE.empty:
            return jsonify({"error": "No working table data"}), 400

        df = WORKING_TABLE.copy()
        # Filter by user-specified time range
        dt = pd.to_datetime(df["Timestamp"], format="%d/%m/%Y %H:%M:%S", errors="coerce")
        numeric = (dt.astype(np.int64) // 1_000_000)
        df["__ms__"] = numeric

        if start_ms is not None and end_ms is not None:
            df = df[(df["__ms__"] >= start_ms) & (df["__ms__"] <= end_ms)]

        df.drop(columns=["__ms__"], inplace=True, errors="ignore")
        if df.empty:
            return jsonify({"error": "No data in that range"}), 400

        cols = df.columns.tolist()
        # reorder so Timestamp is first
        if "Timestamp" in cols:
            cols.remove("Timestamp")
            final_cols = ["Timestamp"] + cols
        else:
            final_cols = cols

        df_export = df[final_cols].copy()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"

        # Write header
        for i, cName in enumerate(final_cols, start=1):
            ws.cell(row=1, column=i, value=cName).alignment = Alignment(horizontal="center", vertical="center")

        # Write data
        for row_idx in range(len(df_export)):
            row_data = df_export.iloc[row_idx]
            for col_idx, cName in enumerate(final_cols, start=1):
                val = row_data[cName]
                ws.cell(row=row_idx+2, column=col_idx, value=val).alignment = Alignment(
                    horizontal="center", vertical="center"
                )

        # Adjust column widths
        for i, cName in enumerate(final_cols, start=1):
            length = len(str(cName))
            col_values = df_export[cName].astype(str)
            max_val_len = col_values.map(len).max() if not col_values.empty else 0
            if max_val_len > length:
                length = max_val_len
            ws.column_dimensions[get_column_letter(i)].width = length + 2

        now = datetime.datetime.now()
        ds = now.strftime("%Y%m%d")
        fname = f"FH {fnum} {bname} {ds}.xlsx"
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        r = make_response(bio.read())
        r.headers["Content-Disposition"] = f'attachment; filename="{fname}"'
        r.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        python_logger.info(f"Excel export success: {fname}")
        return r
    except Exception as e:
        python_logger.error(f"Excel export error: {e}")
        return jsonify({"error": str(e)}), 500

###############################################################################
# EXPORT TO CSV
###############################################################################
@app.route("/export_csv", methods=["POST"])
def export_csv():
    global WORKING_TABLE
    try:
        req = request.get_json()
        start_ms = req.get("startDateUnixMillis")
        end_ms = req.get("endDateUnixMillis")
        bname = req.get("bargeName", "UnknownBarge")
        fnum = req.get("fhNumber", "0000")

        if WORKING_TABLE is None or WORKING_TABLE.empty:
            return jsonify({"error": "No working table data"}), 400

        df = WORKING_TABLE.copy()
        dt = pd.to_datetime(df["Timestamp"], format="%d/%m/%Y %H:%M:%S", errors="coerce")
        numeric = (dt.astype(np.int64) // 1_000_000)
        df["__ms__"] = numeric
        if start_ms is not None and end_ms is not None:
            df = df[(df["__ms__"] >= start_ms) & (df["__ms__"] <= end_ms)]
        df.drop(columns=["__ms__"], inplace=True, errors="ignore")
        if df.empty:
            return jsonify({"error": "No data in that range"}), 400

        cols = df.columns.tolist()
        if "Timestamp" in cols:
            cols.remove("Timestamp")
            final_cols = ["Timestamp"] + cols
        else:
            final_cols = cols
        df_export = df[final_cols].copy()

        output = io.StringIO()
        writer = csv.writer(output, delimiter=",", quotechar='"', quoting=csv.QUOTE_MINIMAL)

        writer.writerow(final_cols)  # header
        for i in range(len(df_export)):
            row_data = df_export.iloc[i].tolist()
            row_str = [str(x) if pd.notna(x) else "" for x in row_data]
            writer.writerow(row_str)

        csv_data = output.getvalue()
        output.close()
        now = datetime.datetime.now()
        ds = now.strftime("%Y%m%d")
        fname = f"FH_{fnum}_{bname}_{ds}.csv"
        resp = make_response(csv_data)
        resp.headers["Content-Disposition"] = f'attachment; filename="{fname}"'
        resp.headers["Content-Type"] = "text/csv"
        return resp
    except Exception as e:
        python_logger.error(f"CSV export error: {e}")
        return jsonify({"error": str(e)}), 500

###############################################################################
# SITE SETTINGS
###############################################################################
@app.route("/site_settings", methods=["GET","POST"])
def site_settings():
    sp = get_site_settings_path()
    if request.method=="GET":
        if os.path.exists(sp):
            with open(sp,"r") as f:
                d = json.load(f)
            # Ensure default values for any missing fields
            if "darkMode" not in d: d["darkMode"] = False
            if "sortOrder" not in d: d["sortOrder"] = "asc"
            if "groupingMode" not in d: d["groupingMode"] = "2"
            if "dataOffset" not in d: d["dataOffset"] = 1
            if "bargeName" not in d: d["bargeName"] = ""
            if "bargeNumber" not in d: d["bargeNumber"] = ""
            if "forwardFill" not in d: d["forwardFill"] = False
            if "pollInterval" not in d: d["pollInterval"] = 5000
            # Default startDate/endDate if not present
            if "startDate" not in d:
                # default to 00:00:00 today
                now = datetime.datetime.now()
                midnight = now.replace(hour=0, minute=0, second=0, microsecond=0)
                d["startDate"] = midnight.strftime("%Y-%m-%d %H:%M:%S")
            if "endDate" not in d:
                # default to now
                d["endDate"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            return jsonify(d)
        else:
            # Default if no file
            now = datetime.datetime.now()
            midnight = now.replace(hour=0, minute=0, second=0, microsecond=0)
            return jsonify({
                "darkMode": False,
                "sortOrder": "asc",
                "groupingMode": "2",
                "dataOffset": 1,
                "bargeName": "",
                "bargeNumber": "",
                "forwardFill": False,
                "pollInterval": 5000,
                "startDate": midnight.strftime("%Y-%m-%d %H:%M:%S"),
                "endDate": now.strftime("%Y-%m-%d %H:%M:%S")
            })
    else:
        try:
            d = request.get_json()
            atomic_write_json(sp, d)
            return jsonify({"status":"ok"})
        except:
            return jsonify({"error":"fail"}),500

###############################################################################
# TAG SETTINGS
###############################################################################
@app.route("/tag_settings", methods=["GET","POST"])
def tag_settings():
    tp = get_tag_settings_path()
    if request.method=="GET":
        if os.path.exists(tp):
            with open(tp, "r") as f:
                return jsonify(json.load(f))
        return jsonify({"scale_factors":{},"max_decimal":{},"error_value":{}})
    else:
        try:
            d = request.get_json()
            atomic_write_json(tp, d)
            return jsonify({"status":"ok"})
        except:
            return jsonify({"error":"fail"}),500

###############################################################################
# CLEAR CACHE, SHUTDOWN, RESTART
###############################################################################
@app.route("/clear_cache", methods=["POST"])
def clear_cache():
    global RAW_TABLE, WORKING_TABLE, TAGLIST_CACHE, TAG_COVERAGE, RAW_TABLE_SIGNATURE
    with global_lock:
        RAW_TABLE = None
        WORKING_TABLE = None
        TAGLIST_CACHE = None
        TAG_COVERAGE = {}
        RAW_TABLE_SIGNATURE = None
        for path in [
            get_taglist_cache_path(),
            get_raw_table_cache_path(),
            get_working_table_cache_path(),
            get_tag_coverage_cache_path()
        ]:
            if os.path.exists(path):
                try:
                    os.remove(path)
                except:
                    pass
        python_logger.info("Cache cleared via /clear_cache.")
    return jsonify({"status":"cleared"})

@app.route("/shutdown", methods=["POST"])
def shutdown():
    sd = request.environ.get("werkzeug.server.shutdown")
    if sd:
        python_logger.info("Server shutting down via /shutdown endpoint.")
        sd()
    return jsonify({"status":"shutting down"})

@app.route("/restart", methods=["POST"])
def restart():
    def do_restart():
        pyExe = sys.executable
        script = os.path.abspath(__file__)
        time.sleep(1)
        subprocess.Popen([pyExe, script])
        python_logger.info("Server restarting (new process).")
        sys.exit(0)

    threading.Thread(target=do_restart, daemon=True).start()
    return jsonify({"status":"restarting"})

###############################################################################
# SYSTEM TRAY
###############################################################################
def create_image():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    icon_path = os.path.join(base_dir, DATA_DIR, "icon.png")
    if os.path.exists(icon_path):
        return Image.open(icon_path)
    else:
        width, height = 64, 64
        image = Image.new("RGB", (width, height), "black")
        dc = ImageDraw.Draw(image)
        dc.text((10,20), "DE", fill="white")
        return image

def on_open(icon, item):
    webbrowser.open(f"http://127.0.0.1:{UI_PORT}")

def on_restart(icon, item):
    try:
        requests.post(f"http://127.0.0.1:{UI_PORT}/restart")
    except:
        pass

def on_quit(icon, item):
    try:
        requests.post(f"http://127.0.0.1:{UI_PORT}/shutdown")
    except:
        pass
    icon.stop()
    os._exit(0)

def start_tray():
    menu = Menu(
        MenuItem("Open DETool", on_open),
        MenuItem("Restart DETool", on_restart),
        MenuItem("Quit DETool", on_quit)
    )
    ic = pystray.Icon("DETool", create_image(), "DETool", menu)
    ic.run()

###############################################################################
# MAIN
###############################################################################
def run_flask():
    global RAW_TABLE, WORKING_TABLE, TAG_COVERAGE, RAW_TABLE_SIGNATURE

    # Load caches at startup
    RAW_TABLE = load_raw_table_cache()
    WORKING_TABLE = load_working_table_cache()
    TAG_COVERAGE = load_tag_coverage()
    RAW_TABLE_SIGNATURE = get_raw_table_signature(RAW_TABLE)

    app.run(host="127.0.0.1", port=UI_PORT, threaded=True)

if __name__ == "__main__":
    threading.Thread(target=run_flask, daemon=True).start()
    time.sleep(1)
    webbrowser.open(f"http://127.0.0.1:{UI_PORT}")
    start_tray()