import os
import sys
import io
import json
import csv
import time
import math
import threading
import datetime
import subprocess
import requests
import pandas as pd
import numpy as np
import concurrent.futures
from flask import Flask, jsonify, request, send_from_directory, make_response
from threading import Lock
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import pystray
from pystray import Menu, MenuItem
from PIL import Image, ImageDraw
import logging
from logging.handlers import RotatingFileHandler

###############################################################################
# GLOBAL LOCK & APP
###############################################################################
global_lock = Lock()
app = Flask(__name__, static_folder="data")

###############################################################################
# LOGGING
###############################################################################
def ensure_folder(path):
    if not os.path.exists(path):
        os.makedirs(path)
    return path

LOG_BASE = ensure_folder(os.path.join(os.path.expanduser("~"), "Documents", "DETool", "Logs"))

def make_rotating_handler(filename):
    return RotatingFileHandler(
        os.path.join(LOG_BASE, filename),
        mode='a',
        maxBytes=5_000_000,
        backupCount=5,
        delay=0
    )

formatter = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")

python_logger = logging.getLogger("python_exec")
python_logger.setLevel(logging.INFO)
py_handler = make_rotating_handler("python_execution.log")
py_handler.setFormatter(formatter)
python_logger.addHandler(py_handler)

script_logger = logging.getLogger("script_exec")
script_logger.setLevel(logging.INFO)
sc_handler = make_rotating_handler("script_execution.log")
sc_handler.setFormatter(formatter)
script_logger.addHandler(sc_handler)

user_logger = logging.getLogger("user_interactions")
user_logger.setLevel(logging.INFO)
us_handler = make_rotating_handler("user_interactions.log")
us_handler.setFormatter(formatter)
user_logger.addHandler(us_handler)

python_logger.info("DETool server starting...")

###############################################################################
# CONSTANTS
###############################################################################
UI_PORT = 8080
DATA_DIR = "data"

EXTERNAL_TAGLIST_URL = "http://localhost:61185/taglist"
EXTERNAL_VALUES_URL  = "http://localhost:61185/values"

###############################################################################
# GLOBAL DATA
###############################################################################
RAW_TABLE           = None
WORKING_TABLE       = None
TAGLIST_CACHE       = None
TAG_COVERAGE        = {}
LAST_SETTINGS       = None  # site settings
LAST_TAG_SETTINGS   = None  # if you want partial tag logic
RAW_TABLE_SIGNATURE = None

###############################################################################
# PATH HELPERS
###############################################################################
def base_folder():
    p = os.path.join(os.path.expanduser("~"), "Documents", "DETool")
    ensure_folder(p)
    return p

def cache_folder():
    cf= os.path.join(base_folder(), "Cache")
    ensure_folder(cf)
    return cf

def settings_folder():
    sf= os.path.join(base_folder(), "Settings")
    ensure_folder(sf)
    return sf

def get_site_settings_path():
    return os.path.join(settings_folder(), "SiteSettings.json")

def get_tag_settings_path():
    return os.path.join(settings_folder(), "TagSettings.json")

def get_taglist_cache_path():
    return os.path.join(cache_folder(), "Taglist.json")

def get_raw_csv_path():
    return os.path.join(cache_folder(), "RawTable.csv")

def get_working_csv_path():
    return os.path.join(cache_folder(), "WorkingTable.csv")

def get_coverage_json_path():
    return os.path.join(cache_folder(), "TagCoverage.json")

def fmt_timestamp(dt):
    return dt.strftime("%d/%m/%Y %H:%M:%S")

###############################################################################
# CSV LOAD/SAVE FOR DATAFRAMES
###############################################################################
def save_df_to_csv(df: pd.DataFrame, path: str):
    if df is None or df.empty:
        # We'll just create an empty CSV with no rows
        temp = path + ".tmp"
        pd.DataFrame().to_csv(temp, index=False)
        os.replace(temp, path)
        return
    temp= path + ".tmp"
    df.to_csv(temp, index=False)
    os.replace(temp, path)

def load_df_from_csv(path: str) -> pd.DataFrame:
    if not os.path.exists(path):
        return pd.DataFrame()
    try:
        return pd.read_csv(path)
    except Exception as e:
        python_logger.error(f"load_df_from_csv({path}) error: {e}")
        return pd.DataFrame()

###############################################################################
# JSON LOAD/SAVE FOR SETTINGS & COVERAGE
###############################################################################
def atomic_write_json(path, data):
    tmp= path + ".tmp"
    with open(tmp,"w",encoding="utf-8") as f:
        json.dump(data,f)
    os.replace(tmp, path)

def safe_load_json(path, default=None):
    if default is None:
        default={}
    if not os.path.exists(path):
        return default
    try:
        with open(path,"r",encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        python_logger.error(f"safe_load_json({path}) error => {e}")
        return default

###############################################################################
# RAW & WORKING => CSV
###############################################################################
def save_raw_table_cache():
    global RAW_TABLE
    if RAW_TABLE is not None:
        save_df_to_csv(RAW_TABLE, get_raw_csv_path())

def load_raw_table_cache():
    p= get_raw_csv_path()
    if os.path.exists(p):
        df= load_df_from_csv(p)
        if not df.empty:
            python_logger.info("Loaded RAW_TABLE from CSV cache.")
            return df
    return pd.DataFrame(columns=["NumericTimestamp","Timestamp"])

def save_working_table_cache():
    global WORKING_TABLE
    if WORKING_TABLE is not None:
        save_df_to_csv(WORKING_TABLE, get_working_csv_path())

def load_working_table_cache():
    p= get_working_csv_path()
    if os.path.exists(p):
        df= load_df_from_csv(p)
        if not df.empty:
            python_logger.info("Loaded WORKING_TABLE from CSV cache.")
            return df
    return None

###############################################################################
# COVERAGE => JSON
###############################################################################
def save_tag_coverage():
    global TAG_COVERAGE
    atomic_write_json(get_coverage_json_path(), TAG_COVERAGE)

def load_tag_coverage():
    p= get_coverage_json_path()
    if os.path.exists(p):
        try:
            with open(p,"r",encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            python_logger.error(f"load_tag_coverage error => {e}")
    return {}

###############################################################################
# SIGNATURE HELPER
###############################################################################
def get_raw_table_signature(df: pd.DataFrame):
    """Simple signature => (row_count, sorted_cols, last_numeric_ts) to see if changed."""
    if df is None or df.empty or "NumericTimestamp" not in df.columns:
        return (0,(),None)
    row_count= len(df)
    col_tuple= tuple(sorted(df.columns))
    last_ts = df["NumericTimestamp"].iloc[-1]
    return (row_count, col_tuple, last_ts)

###############################################################################
# LOG EVENT
###############################################################################
@app.route("/log_event", methods=["POST"])
def log_event():
    try:
        d=request.get_json()
        msg_type= d.get("type","script").lower()
        msg= d.get("message","")
        if msg_type=="user":
            user_logger.info(msg)
        else:
            script_logger.info(msg)
        return jsonify({"status":"logged"})
    except Exception as e:
        python_logger.error(f"log_event error => {e}")
        return jsonify({"error":str(e)}),500

###############################################################################
# STATIC
###############################################################################
@app.route("/")
def root():
    return send_from_directory(DATA_DIR, "index.html")

@app.route("/<path:fname>")
def serve_static(fname):
    return send_from_directory(DATA_DIR, fname)

###############################################################################
# SITE SETTINGS
###############################################################################
@app.route("/site_settings", methods=["GET","POST"])
def site_settings():
    sp= get_site_settings_path()
    if request.method=="GET":
        d= safe_load_json(sp,{
            "darkMode":False,
            "sortOrder":"asc",
            "groupingMode":"2",
            "dataOffset":1,
            "bargeName":"",
            "bargeNumber":"",
            "forwardFill":False,
            "pollInterval":5000,
            "startDate":"",
            "endDate":""
        })
        # ensure default start/end
        if not d.get("startDate"):
            now=datetime.datetime.now()
            midnight= now.replace(hour=0,minute=0,second=0,microsecond=0)
            d["startDate"]= midnight.strftime("%Y-%m-%d %H:%M:%S")
        if not d.get("endDate"):
            d["endDate"]= datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        return jsonify(d)
    else:
        newd= request.get_json()
        atomic_write_json(sp,newd)
        return jsonify({"status":"ok"})

###############################################################################
# TAG SETTINGS (If needed)
###############################################################################
@app.route("/tag_settings", methods=["GET","POST"])
def tag_settings():
    path= get_tag_settings_path()
    if request.method=="GET":
        d= safe_load_json(path,{
            "scale_factors":{},
            "error_value":{},
            "max_decimal":{},
            "global_forward_fill":False
        })
        return jsonify(d)
    else:
        newd= request.get_json()
        atomic_write_json(path, newd)
        return jsonify({"status":"ok"})

###############################################################################
# TAGLIST
###############################################################################
@app.route("/taglist")
def taglist():
    global TAGLIST_CACHE
    refresh= request.args.get("refresh","false").lower() in ["true","1"]
    cpath= get_taglist_cache_path()

    if not refresh and TAGLIST_CACHE:
        return jsonify(TAGLIST_CACHE)
    if not refresh and os.path.exists(cpath):
        try:
            with open(cpath,"r") as f:
                data=json.load(f)
            TAGLIST_CACHE=data
            return jsonify(data)
        except:
            pass
    # fetch external
    try:
        r= requests.get(EXTERNAL_TAGLIST_URL, timeout=15)
        r.raise_for_status()
        data= r.json()
        atomic_write_json(cpath, data)
        TAGLIST_CACHE= data
        return jsonify(data)
    except:
        if TAGLIST_CACHE:
            return jsonify(TAGLIST_CACHE)
        return jsonify([])

###############################################################################
# FETCH EXTERNAL
###############################################################################
def fetch_values(tag, start_sec, end_sec):
    """Calls external endpoint for a single tag, given startSec and endSec, returning (tag, dataList)."""
    try:
        r= requests.get(
            EXTERNAL_VALUES_URL,
            params={"tag":tag,"startDateUnixSeconds":start_sec,"endDateUnixSeconds":end_sec},
            timeout=15
        )
        r.raise_for_status()
        return tag, r.json()
    except Exception as ex:
        python_logger.error(f"fetch_values => {tag} {start_sec}-{end_sec}: {ex}")
        return tag, []

###############################################################################
# UNION INTERVALS
###############################################################################
def union_intervals(ivlist):
    if not ivlist:
        return []
    srt= sorted(ivlist, key=lambda x:x[0])
    merged=[ srt[0] ]
    for i in range(1,len(srt)):
        cur=srt[i]
        last= merged[-1]
        if cur[0]<= last[1]:
            merged[-1]= (last[0], max(last[1], cur[1]))
        else:
            merged.append(cur)
    return merged

def remove_tag_coverage(tag):
    global TAG_COVERAGE, RAW_TABLE
    if tag in TAG_COVERAGE:
        del TAG_COVERAGE[tag]
    if RAW_TABLE is not None and not RAW_TABLE.empty and tag in RAW_TABLE.columns:
        RAW_TABLE.drop(columns=[tag], inplace=True, errors="ignore")

###############################################################################
# MERGE => RAW_TABLE
###############################################################################
def merge_new_data_into_raw_table(df_new):
    """df_new => columns [NumericTimestamp, Timestamp, tagName]."""
    global RAW_TABLE
    if RAW_TABLE is None or RAW_TABLE.empty:
        RAW_TABLE= df_new.sort_values("NumericTimestamp").reset_index(drop=True)
        return
    mg= pd.merge(
        RAW_TABLE, df_new,
        on="NumericTimestamp", how="outer", suffixes=("","_new")
    )
    for c in df_new.columns:
        if c in ["Timestamp","NumericTimestamp"]:
            continue
        newc= c+"_new"
        if newc in mg.columns:
            mg[c] = mg[c].where(mg[newc].isna(), mg[newc])
            mg.drop(columns=[newc], inplace=True, errors="ignore")
    if "Timestamp_new" in mg.columns:
        mg.drop(columns=["Timestamp_new"], inplace=True)
    mg["Timestamp"]= mg["NumericTimestamp"].apply(
        lambda x: fmt_timestamp(pd.to_datetime(x, unit="ms"))
    )
    mg.sort_values("NumericTimestamp", inplace=True)
    mg.reset_index(drop=True, inplace=True)
    RAW_TABLE= mg

###############################################################################
# FETCH_DATA
###############################################################################
@app.route("/fetch_data", methods=["POST"])
def fetch_data_endpoint():
    global RAW_TABLE, TAG_COVERAGE, RAW_TABLE_SIGNATURE
    body= request.get_json()
    if not body:
        return jsonify({"error":"Missing JSON"}),400

    tags= body.get("tags",[])
    st  = body.get("startDateUnixSeconds")
    en  = body.get("endDateUnixSeconds")
    if not tags or st is None or en is None:
        return jsonify({"error":"Missing fields"}),400

    old_sig= RAW_TABLE_SIGNATURE
    data_changed= False

    with global_lock:
        if RAW_TABLE is None:
            RAW_TABLE= pd.DataFrame(columns=["NumericTimestamp","Timestamp"])

        old_tags= set(TAG_COVERAGE.keys())
        new_tags= set(tags)
        removed= old_tags- new_tags
        for rt in removed:
            remove_tag_coverage(rt)

        futs=[]
        with concurrent.futures.ThreadPoolExecutor(max_workers=min(len(tags),4)) as exe:
            for tg in tags:
                if tg not in TAG_COVERAGE:
                    TAG_COVERAGE[tg]=[]
                coverage= union_intervals(TAG_COVERAGE[tg])
                cS,cE= st,en
                missing=[]
                for (cvS,cvE) in coverage:
                    if cvE<cS or cvS>cE:
                        continue
                    if cvS> cS:
                        missing.append((cS, min(cvS,cE)))
                    if cvE> cS:
                        cS= max(cS, cvE)
                    if cS> cE:
                        break
                if cS< cE:
                    missing.append((cS,cE))

                for (miS,miE) in missing:
                    if miE<=miS: continue
                    fu= exe.submit(fetch_values, tg, miS, miE)
                    futs.append((fu,tg,miS,miE))

        for (fu,tg,fs,fe) in futs:
            try:
                tgFetched, arr= fu.result()
                if not arr: continue
                df= pd.DataFrame(arr)
                if df.empty: continue
                df["Value"]= pd.to_numeric(df["Value"], errors="coerce")
                df.replace([np.inf,-np.inf], np.nan, inplace=True)
                df["Timestamp"]= pd.to_datetime(df["Date"], errors="coerce")
                df["NumericTimestamp"]= (df["Timestamp"].astype(np.int64)//1_000_000)
                df.sort_values("Timestamp", inplace=True)
                df_ren= df[["NumericTimestamp","Timestamp","Value"]].rename(columns={"Value":tg})
                merge_new_data_into_raw_table(df_ren)
                TAG_COVERAGE[tg].append((fs,fe))
                TAG_COVERAGE[tg]= union_intervals(TAG_COVERAGE[tg])
            except Exception as e:
                python_logger.error(f"partial fetch error => {tg} => {e}")

        new_sig= get_raw_table_signature(RAW_TABLE)
        data_changed= (new_sig!= old_sig)
        if data_changed:
            RAW_TABLE_SIGNATURE= new_sig
            save_raw_table_cache()
            save_tag_coverage()

    return jsonify({"status":"ok", "newData": data_changed, "redrawNeeded": data_changed})

###############################################################################
# BUILD WORKING TABLE
###############################################################################
@app.route("/build_working_table", methods=["POST"])
def build_working_table_endpoint():
    global RAW_TABLE, WORKING_TABLE, RAW_TABLE_SIGNATURE
    global LAST_SETTINGS

    if RAW_TABLE is None or RAW_TABLE.empty:
        return jsonify({"data":[],"redrawNeeded":False})

    req= request.get_json()
    new_offset= float(req.get("dataOffset", 1))
    new_ff    = bool(req.get("forwardFill", False))

    # load site settings from disk, if you want to unify it. 
    # Or directly trust these from the request
    need_rebuild= False
    if LAST_SETTINGS is None:
        need_rebuild= True
    else:
        old_off= float(LAST_SETTINGS.get("dataOffset",1))
        old_ff = bool(LAST_SETTINGS.get("forwardFill",False))
        if old_off!= new_offset or old_ff!= new_ff:
            need_rebuild= True

    old_len= len(WORKING_TABLE) if WORKING_TABLE is not None else 0
    raw_len= len(RAW_TABLE)
    if raw_len!= old_len:
        need_rebuild= True

    if need_rebuild:
        user_logger.info(f"Rebuilding WORKING_TABLE offset={new_offset}, ff={new_ff}")
        with global_lock:
            WORKING_TABLE= do_build_working_table(RAW_TABLE, new_offset, new_ff)
            LAST_SETTINGS= {"dataOffset": new_offset, "forwardFill": new_ff}
            save_working_table_cache()
    else:
        python_logger.info("No rebuild needed for WORKING_TABLE")

    if WORKING_TABLE is None:
        return jsonify({"data":[],"redrawNeeded":need_rebuild})

    df_s= WORKING_TABLE.replace([np.inf,-np.inf,np.nan], None)
    return jsonify({"data": df_s.to_dict(orient="records"), "redrawNeeded": need_rebuild})

def do_build_working_table(raw_df: pd.DataFrame, offset_hours: float, forward_fill: bool):
    if raw_df is None or raw_df.empty:
        return None

    df= raw_df.copy()
    # forward fill if needed
    if forward_fill:
        df= df.ffill()

    # offset
    offMs= int(offset_hours*3600000)
    df["NumericTimestamp"]= df["NumericTimestamp"] + offMs
    df["Timestamp"]= df["NumericTimestamp"].apply(
        lambda x: fmt_timestamp(pd.to_datetime(x, unit="ms"))
    )

    return df

###############################################################################
# MULTI-ROW EXCEL
###############################################################################
def parse_col_hierarchy(col):
    return col.split(".")

def build_header_rows(columns):
    splitted= [parse_col_hierarchy(c) for c in columns]
    if not splitted:
        return [[]]
    max_depth= max(len(sp) for sp in splitted)
    # rowHeaders[rIndex][colIndex]
    rowHeaders= [ [""]*len(columns) for _ in range(max_depth) ]
    for cIndex, spList in enumerate(splitted):
        for rIndex, token in enumerate(spList):
            rowHeaders[rIndex][cIndex]= token
    return rowHeaders

def apply_header_merges(ws, rowData, startRow):
    """For consecutive duplicates in each row, merge them."""
    for rIndex, rowArr in enumerate(rowData):
        rowNum= startRow + rIndex
        lastVal= None
        spanStart= 1
        for i,val in enumerate(rowArr):
            colPos= i+1
            if val!= lastVal:
                if lastVal is not None:
                    if colPos-1> spanStart:
                        ws.merge_cells(start_row=rowNum, start_column=spanStart,
                                       end_row=rowNum, end_column=(colPos-1))
                spanStart= colPos
                lastVal= val
        if lastVal is not None and spanStart< len(rowArr):
            if len(rowArr)> spanStart:
                ws.merge_cells(start_row=rowNum, start_column=spanStart,
                               end_row=rowNum, end_column=len(rowArr))

@app.route("/export_excel", methods=["POST"])
def export_excel():
    global WORKING_TABLE
    try:
        body= request.get_json()
        start_ms= body.get("startDateUnixMillis")
        end_ms  = body.get("endDateUnixMillis")
        bname   = body.get("bargeName","UnknownBarge")
        fnum    = body.get("fhNumber","0000")

        if WORKING_TABLE is None or WORKING_TABLE.empty:
            return jsonify({"error":"No working table data"}),400

        df= WORKING_TABLE.copy()
        dt= pd.to_datetime(df["Timestamp"],format="%d/%m/%Y %H:%M:%S",errors="coerce")
        ms= dt.astype(np.int64)//1_000_000
        df["__ms__"]= ms
        if start_ms is not None and end_ms is not None:
            df= df[(df["__ms__"]>= start_ms)&(df["__ms__"]<= end_ms)]
        df.drop(columns=["__ms__"], inplace=True, errors="ignore")
        if df.empty:
            return jsonify({"error":"No data in that range"}),400

        cols= df.columns.tolist()
        if "Timestamp" in cols:
            cols.remove("Timestamp")
            cols=["Timestamp"]+ cols
        df_exp= df[cols].copy()

        wb= openpyxl.Workbook()
        ws= wb.active
        ws.title= "Data"

        # big top row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
        topCell= ws.cell(row=1, column=1, value="DETool Multi-row Export")
        topCell.alignment= Alignment(horizontal="center", vertical="center")

        # multi-row header
        rowHeaders= build_header_rows(cols)
        for rIndex, rowArr in enumerate(rowHeaders):
            rowNum= 2+ rIndex
            for cIndex, val in enumerate(rowArr):
                cell= ws.cell(row=rowNum, column=cIndex+1, value= val)
                cell.alignment= Alignment(horizontal="center",vertical="center")
        apply_header_merges(ws, rowHeaders, 2)

        dataStart= 2+ len(rowHeaders)
        for row_i in range(len(df_exp)):
            rowDat= df_exp.iloc[row_i]
            for col_i, colName in enumerate(cols):
                val= rowDat[colName]
                ws.cell(row= dataStart+row_i, column= col_i+1, value= val).alignment= Alignment(horizontal="center", vertical="center")

        # auto-size
        for i, cName in enumerate(cols, start=1):
            length= len(str(cName))
            colVals= df_exp[cName].astype(str)
            maxLen= colVals.map(len).max() if not colVals.empty else 0
            length= max(length, maxLen)
            ws.column_dimensions[get_column_letter(i)].width= length+2

        now= datetime.datetime.now()
        ds= now.strftime("%Y%m%d")
        fname= f"FH {fnum} {bname} {ds}.xlsx"
        bio= io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        resp= make_response(bio.read())
        resp.headers["Content-Disposition"]= f'attachment; filename="{fname}"'
        resp.headers["Content-Type"]= "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        python_logger.info(f"Excel export => {fname}")
        return resp
    except Exception as e:
        python_logger.error(f"Excel export error => {e}")
        return jsonify({"error":str(e)}),500

###############################################################################
# CLEAR CACHE
###############################################################################
@app.route("/clear_cache", methods=["POST"])
def clear_cache():
    global RAW_TABLE, WORKING_TABLE, TAGLIST_CACHE, TAG_COVERAGE, RAW_TABLE_SIGNATURE
    global LAST_SETTINGS, LAST_TAG_SETTINGS

    with global_lock:
        RAW_TABLE=None
        WORKING_TABLE=None
        TAGLIST_CACHE=None
        TAG_COVERAGE={}
        RAW_TABLE_SIGNATURE=None
        LAST_SETTINGS=None
        LAST_TAG_SETTINGS=None

        # remove cached files
        for p in [ get_taglist_cache_path(),
                    get_raw_csv_path(),
                    get_working_csv_path(),
                    get_coverage_json_path() ]:
            if os.path.exists(p):
                try: os.remove(p)
                except: pass

        python_logger.info("Cache cleared via /clear_cache.")
    return jsonify({"status":"cleared"})

###############################################################################
# SHUTDOWN / RESTART
###############################################################################
@app.route("/shutdown", methods=["POST"])
def shutdown():
    fn= request.environ.get("werkzeug.server.shutdown")
    if fn:
        python_logger.info("Shutdown via /shutdown.")
        fn()
    return jsonify({"status":"shutting down"})

@app.route("/restart", methods=["POST"])
def restart():
    def do_restart():
        py= sys.executable
        sc= os.path.abspath(__file__)
        time.sleep(1)
        subprocess.Popen([py,sc])
        python_logger.info("Server restarting (new process).")
        sys.exit(0)
    threading.Thread(target=do_restart, daemon=True).start()
    return jsonify({"status":"restarting"})

###############################################################################
# SYSTEM TRAY
###############################################################################
def create_image():
    base= os.path.dirname(os.path.abspath(__file__))
    ipath= os.path.join(base, DATA_DIR,"icon.png")
    if os.path.exists(ipath):
        return Image.open(ipath)
    # fallback
    w,h=64,64
    im= Image.new("RGB",(w,h),"black")
    dc= ImageDraw.Draw(im)
    dc.text((10,20),"DE",fill="white")
    return im

def on_open(icon,item):
    import webbrowser
    webbrowser.open(f"http://127.0.0.1:{UI_PORT}")

def on_restart(icon,item):
    try:
        requests.post(f"http://127.0.0.1:{UI_PORT}/restart")
    except: pass

def on_quit(icon,item):
    try:
        requests.post(f"http://127.0.0.1:{UI_PORT}/shutdown")
    except: pass
    icon.stop()
    os._exit(0)

def start_tray():
    menu= Menu(
        MenuItem("Open DETool", on_open),
        MenuItem("Restart DETool", on_restart),
        MenuItem("Quit DETool", on_quit)
    )
    ic= pystray.Icon("DETool", create_image(), "DETool", menu)
    ic.run()

###############################################################################
# MAIN
###############################################################################
def run_flask():
    global RAW_TABLE, WORKING_TABLE, TAG_COVERAGE, RAW_TABLE_SIGNATURE
    RAW_TABLE= load_raw_table_cache()
    WORKING_TABLE= load_working_table_cache()
    TAG_COVERAGE= load_tag_coverage()
    RAW_TABLE_SIGNATURE= get_raw_table_signature(RAW_TABLE)
    app.run(host="127.0.0.1", port=UI_PORT, threaded=True)

if __name__=="__main__":
    threading.Thread(target=run_flask,daemon=True).start()
    time.sleep(1)
    import webbrowser
    webbrowser.open(f"http://127.0.0.1:{UI_PORT}")
    start_tray()
